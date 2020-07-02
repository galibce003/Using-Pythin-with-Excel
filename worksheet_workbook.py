from openpyxl.workbook import Workbook
from openpyxl import load_workbook

wb = Workbook()    #define Workbook.
ws = wb.active     #define Worksheet.

ws1= wb.create_sheet("New1")    #Create a new Sheet.
ws2= wb.create_sheet("New2",0)  #Here 0 is the index. It will always come 1st.
ws.title = "Main"               #Rename the active worksheet

print(wb.sheetnames)            #It shows the sheets



#Example
wb2 = load_workbook("C:/Users/Mehedi Hassan Galib/Desktop/Python/ggg.xlsx")
new_sheet = wb2.create_sheet("New")
active_sheet = wb2.active
cell = active_sheet['A1']
print(cell.value)    #RETURN THE VALUE OF CELL A1

active_sheet["A1"] = "Time"  #Change the value of cell A1
wb2.save("modified2.xlsx")   #Create a new excel file
